import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 添加项目根目录到 Python 路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(project_root)

from utils.mysqldb import MysqlDB

def fetch_data():
    """从数据库获取数据"""
    try:
        # 数据库配置
        db_config = {
            'ip': 'localhost',
            'port': 3306,
            'db': 'wechat',
            'user': 'root',
            'passwd': ''
        }
        db = MysqlDB(**db_config)
        
        # 查询数据
        sql = """
            SELECT 
                title as '文章标题',
                publish_time as '发布时间',
                read_num as '阅读量',
                like_num as '点赞数',
                create_time as '抓取时间'
            FROM wechat_topic 
            ORDER BY create_time DESC;
        """
        data = db.query(sql)
        return data
    except Exception as e:
        print(f"获取数据��败: {e}")
        return None

def create_excel(data):
    """创建Excel文件"""
    try:
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "微信文章数据"

        # 设置表头
        headers = ['文章标题', '发布时间', '阅读量', '点赞数', '抓取时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # 设置表头样式
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # 调整列宽
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            # 获取该列所有单元格
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置列宽（最小10，最大50）
            adjusted_width = min(max(10, max_length + 2), 50)
            ws.column_dimensions[column].width = adjusted_width

        # 保存文件
        output_file = '微信文章数据.xlsx'
        wb.save(output_file)
        print(f"\nExcel文件已保存: {output_file}")
        return True
    except Exception as e:
        print(f"创建Excel文件失败: {e}")
        return False

def main():
    """主函数"""
    print("\n开始导出数据到Excel...")
    
    # 获取数据
    data = fetch_data()
    if not data:
        print("没有找到数据，退出程序")
        return
    
    print(f"成功获取 {len(data)} 条数据")
    
    # 创建Excel文件
    if create_excel(data):
        print("数据导出完成")
    else:
        print("数据导出失败")

if __name__ == "__main__":
    main() 